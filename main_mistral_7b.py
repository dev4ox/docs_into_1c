import os
import json
import pandas as pd
from pathlib import Path
import importlib.util
from llama_cpp import Llama
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Настройки хранятся в той же папке, что и исполняемый файл
BASE_DIR = Path(__file__).resolve().parent
SETTINGS_PATH = BASE_DIR / "settings.py"
spec = importlib.util.spec_from_file_location("settings", SETTINGS_PATH)
settings = importlib.util.module_from_spec(spec)
spec.loader.exec_module(settings)

# Объект модели Llama с поддержкой CUDA (gguf)
llm = Llama(
    model_path="models/Mistral-7B-Instruct-v0.3-Q4_K_M.gguf",
    n_gpu_layers=-1,
    n_ctx=2048
)

# Новый промпт, дающий четкие инструкции и примеры
initial_prompt = """Ты эксперт по извлечению данных из технических описаний светильников.
Твоя задача – проанализировать входной текст и извлечь параметры для заполнения таблицы "Форма 2".
Выводи ровно один JSON-словарь с ключами (значения выводи как строки):
"Номенклатура", "Мощность, Вт", "Св. поток, Лм", "IP", "Длина, мм", "Ширина, мм", "Высота, мм", "Габариты", "Рассеиватель", "Цвет. температура, К", "Вес, кг", "Напряжение, В", "Срок службы (работы) светильника", "Температура эксплуатации", "Материал корпуса", "Тип", "Тип КСС", "Род тока", "Гарантия", "Индекс цветопередачи (CRI, Ra)", "Класс защиты от поражения электрическим током", "Коэффициент мощности (Pf)", "С регулятором яркости (диммирование)", "Ударопрочность", "Класс взрывозащиты (Ex)", "Класс пожароопасности", "Цвет корпуса", "Коэффициент пульсаций", "Прочее".

Параметры могут быть заданы с диапазонами или с квалификаторами (например, "не более", "не менее", "от X до Y", "±10", "+-10", "около"). Если параметр отсутствует или его значение не распознано, верни "не указано". Если есть дополнительные характеристики, не подходящие под стандартные поля, помести их в поле "Прочее".

Пример:
Входной текст:
"Светильник светодиодный CSVT Айсберг-38 или эквивалент. Вид крепления - накладной/подвесной. Тип - настенные/потолочные. Мощность светильника не более 38 Вт. Тип лампы - встроенные светодиоды. Материал корпуса/плафона/арматуры - пластик ABS/полимерный пластик SAN. Цвет плафона/арматуры - белый/серый. Напряжение питания 220 В. Длина светильника 1270 мм. Ширина светильника 152 мм (отклонение в размерах не более ± 10 мм). Световой поток – не менее 4800 Лм. Защита от пыли и влаги – не хуже IP65. Цветовая температура - 5000 К. Цветопередача 80 Ra. Влагозащищенный - да. Пылезащищенный - да. Форма корпуса - прямоугольник."
Вывод:
{
  "Номенклатура": "Светильник светодиодный CSVT Айсберг-38 или эквивалент",
  "Мощность, Вт": "не более 38 Вт",
  "Св. поток, Лм": "не менее 4800 Лм",
  "IP": "не хуже IP65",
  "Длина, мм": "1270 мм",
  "Ширина, мм": "152 мм",
  "Высота, мм": "не указано",
  "Габариты": "1270 мм x 152 мм (отклонение в размерах не более ± 10 мм)",
  "Рассеиватель": "не указано",
  "Цвет. температура, К": "5000 К",
  "Вес, кг": "не указано",
  "Напряжение, В": "220 В",
  "Срок службы (работы) светильника": "не указано",
  "Температура эксплуатации": "не указано",
  "Материал корпуса": "пластик ABS/полимерный пластик SAN",
  "Тип": "накладной/подвесной; настенные/потолочные",
  "Тип КСС": "не указано",
  "Род тока": "не указано",
  "Гарантия": "не указано",
  "Индекс цветопередачи (CRI, Ra)": "80 Ra",
  "Класс защиты от поражения электрическим током": "не указано",
  "Коэффициент мощности (Pf)": "не указано",
  "С регулятором яркости (диммирование)": "не указано",
  "Ударопрочность": "не указано",
  "Класс взрывозащиты (Ex)": "не указано",
  "Класс пожароопасности": "не указано",
  "Цвет корпуса": "белый/серый",
  "Коэффициент пульсаций": "не указано",
  "Прочее": "Тип лампы - встроенные светодиоды; (отклонение в размерах не более ± 10 мм); Влагозащищенный - да; Пылезащищенный - да; Форма корпуса - прямоугольник"
}
Return only the JSON object.
"""

def extract_with_llama2(text):
    prompt = initial_prompt + "\n\nText:\n" + text + "\n\nJSON:"
    output = llm(prompt=prompt, max_tokens=512, temperature=0.0)
    result_text = output["choices"][0]["text"].strip()
    try:
        data = json.loads(result_text)
    except Exception as e:
        print("Error parsing JSON:", e)
        print("Raw output:", result_text)
        data = {}
    return data

def append_df_to_excel(filename, df, sheet_name='Sheet1'):
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append(row)
        wb.save(filename)
    else:
        df.to_excel(filename, index=False, sheet_name=sheet_name)

class UnifiedExcelParser:
    PRODUCT_NAMES = settings.product_names
    def __init__(self, file_path):
        self.file_path = Path(file_path)
    def is_product_name(self, text):
        return any(text.lower().startswith(name.lower()) for name in self.PRODUCT_NAMES)
    def detect_engine(self):
        ext = self.file_path.suffix.lower()
        if ext in ['.xlsx', '.xlsm']:
            return 'openpyxl'
        elif ext == '.xls':
            return 'xlrd'
        else:
            return 'openpyxl'
    def parse_excel(self):
        if not self.file_path.exists():
            print(f"File not found: {self.file_path}")
            return []
        engine = self.detect_engine()
        df = pd.read_excel(self.file_path, header=None, engine=engine)
        products = []
        if df.shape[1] == 1:
            products = self.parse_single_column(df)
        else:
            name_column = None
            for col in df.columns:
                for row_idx in range(min(15, len(df))):
                    cell_value = str(df.iloc[row_idx, col]).lower() if pd.notna(df.iloc[row_idx, col]) else ""
                    if "наименование" in cell_value:
                        name_column = col
                        break
                if name_column is not None:
                    break
            if name_column is None:
                products = self.parse_single_column(df)
            else:
                products = self.parse_multi_column(df, name_column, (name_column + 2 < df.shape[1]))
        return products
    def parse_single_column(self, df):
        products = []
        current_text = ""
        for index, row in df.iterrows():
            cell_value = str(row[0]).strip()
            if self.is_product_name(cell_value):
                if current_text:
                    products.append(current_text)
                current_text = cell_value
            else:
                current_text += " " + cell_value
        if current_text:
            products.append(current_text)
        return products
    def parse_multi_column(self, df, name_column, extra_char):
        products = []
        current_text = ""
        for index, row in df.iterrows():
            if pd.isna(row[name_column]):
                continue
            cell_value = str(row[name_column]).strip().replace("\t", " ")
            char_value = ""
            if extra_char:
                if name_column + 1 < df.shape[1]:
                    char_value += " " + str(row[name_column + 1]).strip().replace("\t", " ").replace("\n", " ")
                if name_column + 2 < df.shape[1]:
                    char_value += " " + str(row[name_column + 2]).strip().replace("\t", " ").replace("\n", " ")
            else:
                if name_column + 1 < df.shape[1]:
                    char_value = " " + str(row[name_column + 1]).strip().replace("\t", " ").replace("\n", " ")
            if self.is_product_name(cell_value):
                if current_text:
                    products.append(current_text)
                current_text = cell_value + char_value
            else:
                current_text += " " + cell_value + char_value
        if current_text:
            products.append(current_text)
        return products

if __name__ == "__main__":
    input_file_path = Path("test_data", "input", "ТЗ для 213054.xlsx")
    parser = UnifiedExcelParser(input_file_path)
    product_texts = parser.parse_excel()
    final_columns = [
        "Номенклатура", "Мощность, Вт", "Св. поток, Лм", "IP", "Длина, мм", "Ширина, мм",
        "Высота, мм", "Габариты", "Рассеиватель", "Цвет. температура, К", "Вес, кг",
        "Напряжение, В", "Срок службы (работы) светильника", "Температура эксплуатации",
        "Материал корпуса", "Тип", "Тип КСС", "Род тока", "Гарантия",
        "Индекс цветопередачи (CRI, Ra)", "Класс защиты от поражения электрическим током",
        "Коэффициент мощности (Pf)", "С регулятором яркости (диммирование)",
        "Ударопрочность", "Класс взрывозащиты (Ex)", "Класс пожароопасности",
        "Цвет корпуса", "Коэффициент пульсаций", "Прочее"
    ]
    output_file = "output.xlsx"
    # Если файла output.xlsx нет, создаём его с заголовками.
    if not os.path.exists(output_file):
        pd.DataFrame(columns=final_columns).to_excel(output_file, index=False, sheet_name="Sheet1")
    # Для каждого товара сразу извлекаем данные и добавляем в Excel
    for text in product_texts:
        extracted = extract_with_llama2(text)
        for col in final_columns:
            if col not in extracted:
                extracted[col] = "не указано"
        df_row = pd.DataFrame([extracted], columns=final_columns)
        append_df_to_excel(output_file, df_row, sheet_name="Sheet1")
        print(f"Добавлен товар: {extracted.get('Номенклатура', 'не указано')}")
    print(f"\nДанные успешно добавлены в файл {output_file}.")
