import os
import pandas as pd
from pathlib import Path
import importlib.util
import spacy
from spacy.matcher import PhraseMatcher
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------------- Загрузка настроек из settings.py ----------------
BASE_DIR = Path(__file__).resolve().parent
SETTINGS_PATH = BASE_DIR / "settings.py"

spec = importlib.util.spec_from_file_location("settings", SETTINGS_PATH)
settings = importlib.util.module_from_spec(spec)
spec.loader.exec_module(settings)

# ---------------- Класс для объединённого парсинга Excel ----------------
class UnifiedExcelParser:
    PRODUCT_NAMES = settings.product_names  # список ключевых названий товаров из настроек

    def __init__(self, file_path):
        self.file_path = Path(file_path)
        self.data = []  # сюда будем записывать разобранные товары

    def is_product_name(self, text):
        """Проверяет, начинается ли строка с одного из названий товара."""
        return any(text.lower().startswith(name.lower()) for name in self.PRODUCT_NAMES)

    def detect_engine(self):
        """Определяет движок pd.read_excel по расширению файла."""
        ext = self.file_path.suffix.lower()
        if ext in ['.xlsx', '.xlsm']:
            return 'openpyxl'
        elif ext == '.xls':
            return 'xlrd'
        else:
            return 'openpyxl'

    def parse_excel(self):
        if not self.file_path.exists():
            print(f"File not found: {self.file_path}")
            return
        engine = self.detect_engine()
        print(f"Opening file: {self.file_path} with engine {engine}")
        df = pd.read_excel(self.file_path, header=None, engine=engine)
        # Если в файле только один столбец – используем одностолбцовый разбор
        if df.shape[1] == 1:
            self.parse_single_column(df)
        else:
            # Ищем столбец с ключевым словом "наименование" в первых 15 строках
            name_column = None
            for col in df.columns:
                for row_idx in range(min(15, len(df))):
                    cell_value = str(df.iloc[row_idx, col]).lower() if pd.notna(df.iloc[row_idx, col]) else ""
                    if "наименование" in cell_value:
                        name_column = col
                        print(f"Found 'Наименование' column at index {col}")
                        break
                if name_column is not None:
                    break

            if name_column is None:
                print("Столбец 'Наименование' не найден, пробуем обработку как одностолбцовый файл.")
                self.parse_single_column(df)
            else:
                # Если рядом с найденным столбцом есть ещё ячейки, определяем схему разбора
                extra_char = (name_column + 2 < df.shape[1])
                self.parse_multi_column(df, name_column, extra_char)

    def parse_single_column(self, df):
        """Разбор для файла с одним столбцом (логика варианта 2)."""
        current_product = None
        product_data = {}
        for index, row in df.iterrows():
            cell_value = str(row[0]).strip()
            if self.is_product_name(cell_value):
                if current_product:
                    self.data.append(product_data)
                current_product = cell_value
                product_data = {"0": current_product}
            else:
                product_data[f"{len(product_data)}"] = cell_value
        if current_product:
            self.data.append(product_data)

    def parse_multi_column(self, df, name_column, extra_char):
        """
        Разбор для файлов с несколькими столбцами (варианты 1, 3, 4, 5).
        Если extra_char==True, предполагается, что данные о характеристиках распределены по двум столбцам.
        """
        current_product = None
        product_data = {}
        for index, row in df.iterrows():
            if pd.isna(row[name_column]):
                continue
            cell_value = str(row[name_column]).strip().replace("\t", " ")
            char_value = ""
            if extra_char:
                if name_column + 1 < df.shape[1]:
                    char_value += str(row[name_column + 1]).strip().replace("\t", " ").replace("\n", ";")
                if name_column + 2 < df.shape[1]:
                    char_value += " " + str(row[name_column + 2]).strip().replace("\t", " ").replace("\n", ";")
            else:
                if name_column + 1 < df.shape[1]:
                    char_value = str(row[name_column + 1]).strip().replace("\t", " ").replace("\n", ";")
            if any(name.lower() in cell_value.lower() for name in self.PRODUCT_NAMES):
                if current_product:
                    self.data.append(product_data)
                current_product = cell_value
                product_data = {"0": current_product}
                if char_value:
                    product_data["1"] = char_value
            else:
                product_data[f"{len(product_data)}"] = cell_value
        if current_product:
            self.data.append(product_data)

    def print_data(self):
        if not self.data:
            print("No data parsed!")
        else:
            for product in self.data:
                print(product)

    def process(self):
        if not self.file_path.exists():
            print(f"File not found: {self.file_path}")
            return
        print(f"Processing file: {self.file_path.name}")
        self.parse_excel()
        self.print_data()

# ---------------- Модуль извлечения параметров на основе spaCy ----------------
# Загружаем русскую модель spaCy (убедитесь, что она установлена: python -m spacy download ru_core_news_sm)
nlp = spacy.load("ru_core_news_sm")

# Словарь с синонимами (ключи совпадают с названиями столбцов)
SYNONYMS = {
    "Номенклатура": [
        "серия", "наименование позиции", "наименование", "товар"
    ],
    "Мощность, Вт": [
        "мощность светильника", "энергопотребление", "потребляемая мощность",
        "номинальная мощность", "мощность", "Вт", "W"
    ],
    "Св. поток, Лм": [
        "световой поток", "номинальный световой поток", "Лм", "Lm",
        "общий световой поток модуля светильника"
    ],
    "IP": [
        "ip", "степень защиты", "защита от пыли и влаги", "степень защиты по ГОСТ"
    ],
    "Габариты": [
        "размеры", "габариты", "габаритные размеры",
        "габаритные размеры светильника (без выносных элементов)"
    ],
    "Длина, мм": [
        "длина", "l"
    ],
    "Ширина, мм": [
        "ширина", "b"
    ],
    "Высота, мм": [
        "высота", "h"
    ],
    "Рассеиватель": [
        "рассеиватель", "материал рассеивателя"
    ],
    "Цвет. температура, К": [
        "цветовая температура", "цвет свечения", "температура цвета",
        "коррелированная цветовая температура", "коррелированная цветовая температура по ГОСТ",
        "к"
    ],
    "Вес, кг": [
        "вес", "масса", "общий вес", "нетто", "кг"
    ],
    "Напряжение, В": [
        "входное напряжение", "напряжение", "номинальное напряжение",
        "напряжение питания", "номинальное напряжение питания сети",
        "в", "v"
    ],
    "Температура эксплуатации": [
        "температура", "температура эксплуатации"
    ],
    "Материал корпуса": [
        "материал корпуса", "материал изделия"
    ],
    "Тип": [
        "тип", "тип крепления", "способ монтажа", "назначение",
        "монтаж", "модификация", "установка на", "вид крепления", "тип монтажа светильника"
    ],
    "Срок службы (работы) светильника": [
        "срок службы", "срок работы", "срок эксплуатации", "время работы",
        "эксплуатационный срок", "время службы"
    ],
    "Тип КСС": [
        "тип ксс", "ксс", "линза", "кривая(ые) сила света",
        "кривая", "угол", "тип кривой силы света", "угол излучения",
        "[°]", "светораспределение", "угол светового излучения",
        "класс светораспределения", "угол свечения"
    ],
    "Род тока": [
        "род тока", "тип питания", "род измеряемого тока", "ток"
    ],
    "Гарантия": [
        "гарантия", "гарантийный срок", "срок гарантии",
        "гарантийный срок службы", "гарантийный срок эксплуатации",
        "гарантийный срок работы"
    ],
    "Индекс цветопередачи (CRI, Ra)": [
        "cri", "ra", "индекс цветопередачи, ra", "индекс цветопередачи (cri)",
        "цветовая передача"
    ],
    "Класс защиты от поражения электрическим током": [
        "класс защиты от поражения электрическим током",
        "класс защиты от поражения электрическим током по ГОСТ",
        "класс защиты", "тип защиты от поражения электрическим током"
    ],
    "Цвет корпуса": [
        "цвет корпуса", "цвет изделия", "цвет покраски", "цвет"
    ],
    "Коэффициент пульсаций": [
        "коэффициент пульсаций", "пульсация светового потока", "значение пульсации"
    ]
}

def create_phrase_matcher(nlp, synonyms_dict):
    """
    Создаём общий PhraseMatcher для всех колонок.
    Возвращаем:
    - matcher (PhraseMatcher)
    - mapping (словарь {match_id: column_name})
    """
    matcher = PhraseMatcher(nlp.vocab, attr="LOWER")
    mapping = {}
    for column_name, phrases in synonyms_dict.items():
        patterns = []
        for phrase in phrases:
            phrase_doc = nlp.make_doc(phrase.lower())
            patterns.append(phrase_doc)
        match_id = column_name.lower().replace(" ", "_")
        matcher.add(match_id, patterns)
        mapping[match_id] = column_name
    return matcher, mapping

matcher, mapping = create_phrase_matcher(nlp, SYNONYMS)

def extract_parameters_spacy(text):
    """
    Обрабатывает текст и извлекает параметры согласно словарю синонимов.
    Возвращает словарь вида:
      { "Номенклатура": [...], "Мощность, Вт": [...], ..., "Прочее": [...] }.
    """
    doc = nlp(text)
    extracted = {col: [] for col in SYNONYMS.keys()}
    extracted["Прочее"] = []
    matches = matcher(doc)
    found_spans = {}
    for match_id, start, end in matches:
        col_name = mapping[nlp.vocab.strings[match_id]]
        found_spans[(start, end)] = col_name
    i = 0
    while i < len(doc):
        matched_span = None
        matched_col = None
        span_length = 0
        for (start, end), col_name in found_spans.items():
            if start == i:
                matched_span = doc[start:end].text
                matched_col = col_name
                span_length = end - start
                break
        if matched_span:
            extracted[matched_col].append(matched_span)
            i += span_length
        else:
            token_text = doc[i].text.strip()
            if token_text:
                extracted["Прочее"].append(token_text)
            i += 1
    return extracted

# ---------------- Функция для добавления данных в output.xlsx ----------------
def append_df_to_excel(filename, df, sheet_name='Sheet1'):
    """
    Если файл filename существует, функция находит последнюю заполненную строку
    и добавляет новые данные (без заголовков). Если файла нет, создается новый Excel-файл с заголовками.
    """
    if os.path.exists(filename):
        wb = load_workbook(filename)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append(row)
        wb.save(filename)
    else:
        df.to_excel(filename, index=False, sheet_name=sheet_name)

# ---------------- Основная логика: парсинг, извлечение параметров и запись в output.xlsx ----------------
if __name__ == "__main__":
    # Путь к входящему файлу (например, "ТЗ для GPT.xlsx")
    input_file_path = Path("test_data", "input", "ТЗ для GPT.xlsx")
    parser = UnifiedExcelParser(input_file_path)
    parser.process()  # Разбор Excel-файла и вывод отладочной информации

    # Для каждого товара из Excel объединяем все поля в один текст и извлекаем параметры с помощью spaCy
    filled_forms = []
    for product in parser.data:
        # Объединяем все значения словаря в один текст
        product_text = " ".join(product.values())
        extracted = extract_parameters_spacy(product_text)
        # Преобразуем списки в строки (если список пуст, записываем "не указано")
        form = {}
        for key, values in extracted.items():
            form[key] = "; ".join(values) if values else "не указано"
        # Приоритетное значение "Номенклатура" возьмем из исходного парсера, если оно есть
        if product.get("0"):
            form["Номенклатура"] = product["0"].replace(" или эквивалент.", "").strip()
        filled_forms.append(form)

    # Задаем итоговый порядок столбцов (расширенный список, включающий дополнительные поля)
    final_columns = [
        "Номенклатура",
        "Мощность, Вт",
        "Св. поток, Лм",
        "IP",
        "Габариты",
        "Длина, мм",
        "Ширина, мм",
        "Высота, мм",
        "Рассеиватель",
        "Цвет. температура, К",
        "Вес, кг",
        "Напряжение, В",
        "Температура эксплуатации",
        "Срок службы (работы) светильника",
        "Тип КСС",
        "Род тока",
        "Гарантия",
        "Индекс цветопередачи (CRI, Ra)",
        "Цвет корпуса",
        "Коэффициент пульсаций",
        "Коэффициент мощности (Pf)",
        "Класс взрывозащиты (Ex)",
        "Класс пожароопасности",
        "Класс защиты от поражения электрическим током",
        "Материал корпуса",
        "Тип",
        "Прочее"
    ]
    # Для каждого результата гарантируем наличие всех столбцов
    for form in filled_forms:
        for col in final_columns:
            if col not in form:
                form[col] = "не указано"
    df_form = pd.DataFrame(filled_forms, columns=final_columns)
    print("\nЗаполненная форма:")
    print(df_form.to_string(index=False))

    # Добавляем данные в output.xlsx (новые строки дописываются после существующих)
    output_file = "output.xlsx"
    append_df_to_excel(output_file, df_form, sheet_name="Sheet1")
    print(f"\nДанные успешно добавлены в файл {output_file}.")
