import os
import json
import pandas as pd
from pathlib import Path
import importlib.util
from llama_cpp import Llama
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
# from parsers.pdf import TZ_for_RIR, TZ_for_NIIAR

BASE_DIR = Path(__file__).resolve().parent
SETTINGS_PATH = BASE_DIR / "settings.py"
spec = importlib.util.spec_from_file_location("settings", SETTINGS_PATH)
settings = importlib.util.module_from_spec(spec)
spec.loader.exec_module(settings)

llm = Llama(model_path="models/Qwen2.5-7B-Instruct-1M.gguf",
            n_gpu_layers=-1,
            n_ctx=2048)

# llm = Llama(model_path="models/Qwen2.5-7B-Instruct-1M-GGUF/Qwen2.5-7B-Instruct-1M.gguf",
#             n_ctx=2048)

initial_prompt = """You are an expert data extractor specialized in processing technical descriptions of lighting fixtures.
The input is a single-line text containing all information about one product. The text may include approximate values, ranges, and qualifiers such as "не более", "не менее", "+-", etc.
Extract the following fields and output a valid JSON object with exactly these keys:
"Номенклатура", "Мощность, Вт", "Св. поток, Лм", "IP", "Габариты", "Длина, мм", "Ширина, мм", "Высота, мм", "Рассеиватель", "Цвет. температура, К", "Вес, кг", "Напряжение, В", "Температура эксплуатации", "Срок службы (работы) светильника", "Тип КСС", "Род тока", "Гарантия", "Индекс цветопередачи (CRI, Ra)", "Цвет корпуса", "Коэффициент пульсаций", "Коэффициент мощности (Pf)", "Класс взрывозащиты (Ex)", "Класс пожароопасности", "Класс защиты от поражения электрическим током", "Материал корпуса", "Тип", "Прочее".
For parameters expressed as ranges (e.g. "от X до Y", "X ÷ Y") or with qualifiers ("не более", "не менее"), include the entire expression as found.
If no information is found for a field, output "не указано".
Return only the JSON object.
"""


def extract_with_qwen(text):
    prompt = initial_prompt + "\n\nText:\n" + text + "\n\nExtracted JSON:"
    output = llm(prompt=prompt, max_tokens=512, temperature=0.0)
    result_text = output["choices"][0]["text"].strip()
    try:
        data = json.loads(result_text)
    except Exception as e:
        print("Error parsing JSON:", e)
        print("Raw output:", result_text)
        data = {}
    return data


def append_df_to_excel(filename, df, sheet_name='Sheet1'):
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append(row)
        wb.save(filename)
    else:
        df.to_excel(filename, index=False, sheet_name=sheet_name)


class UnifiedExcelParser:
    PRODUCT_NAMES = settings.product_names

    def __init__(self, file_path):
        self.file_path = Path(file_path)
        self.data = []

    def is_product_name(self, text):
        return any(text.lower().startswith(name.lower()) for name in self.PRODUCT_NAMES)

    def detect_engine(self):
        ext = self.file_path.suffix.lower()
        if ext in ['.xlsx', '.xlsm']:
            return 'openpyxl'
        elif ext == '.xls':
            return 'xlrd'
        else:
            return 'openpyxl'

    def parse_excel(self):
        if not self.file_path.exists():
            print(f"File not found: {self.file_path}")
            return
        engine = self.detect_engine()
        df = pd.read_excel(self.file_path, header=None, engine=engine)
        if df.shape[1] == 1:
            self.parse_single_column(df)
        else:
            name_column = None
            for col in df.columns:
                for row_idx in range(min(15, len(df))):
                    cell_value = str(df.iloc[row_idx, col]).lower() if pd.notna(df.iloc[row_idx, col]) else ""
                    if "наименование" in cell_value:
                        name_column = col
                        break
                if name_column is not None:
                    break
            if name_column is None:
                self.parse_single_column(df)
            else:
                extra_char = (name_column + 2 < df.shape[1])
                self.parse_multi_column(df, name_column, extra_char)

    def parse_single_column(self, df):
        current_text = ""
        for index, row in df.iterrows():
            cell_value = str(row[0]).strip()
            if self.is_product_name(cell_value):
                if current_text:
                    self.data.append({"text": current_text})
                current_text = cell_value
            else:
                current_text += " " + cell_value
        if current_text:
            self.data.append({"text": current_text})

    def parse_multi_column(self, df, name_column, extra_char):
        current_text = ""
        for index, row in df.iterrows():
            if pd.isna(row[name_column]):
                continue
            cell_value = str(row[name_column]).strip().replace("\t", " ")
            char_value = ""
            if extra_char:
                if name_column + 1 < df.shape[1]:
                    char_value += " " + str(row[name_column + 1]).strip().replace("\t", " ").replace("\n", " ")
                if name_column + 2 < df.shape[1]:
                    char_value += " " + str(row[name_column + 2]).strip().replace("\t", " ").replace("\n", " ")
            else:
                if name_column + 1 < df.shape[1]:
                    char_value = " " + str(row[name_column + 1]).strip().replace("\t", " ").replace("\n", " ")
            if any(name.lower() in cell_value.lower() for name in self.PRODUCT_NAMES):
                if current_text:
                    self.data.append({"text": current_text})
                current_text = cell_value + char_value
            else:
                current_text += " " + cell_value + char_value
        if current_text:
            print(current_text)
            self.data.append({"text": current_text})

    def process(self):
        self.parse_excel()


if __name__ == "__main__":
    input_file_path = Path("test_data", "input", "ТЗ для GPT.xlsx")
    parser = UnifiedExcelParser(input_file_path)
    # parser = TZ_for_RIR.StructuredPdfParser(Path("test_data", "input", "ТЗ для РИР.pdf"))
    parser.process()
    filled_forms = []
    final_columns = ["Номенклатура", "Мощность, Вт", "Св. поток, Лм", "IP", "Габариты", "Длина, мм",
                     "Ширина, мм", "Высота, мм", "Рассеиватель", "Цвет. температура, К", "Вес, кг",
                     "Напряжение, В", "Температура эксплуатации", "Срок службы (работы) светильника",
                     "Тип КСС", "Род тока", "Гарантия", "Индекс цветопередачи (CRI, Ra)", "Цвет корпуса",
                     "Коэффициент пульсаций", "Коэффициент мощности (Pf)", "Класс взрывозащиты (Ex)",
                     "Класс пожароопасности", "Класс защиты от поражения электрическим током",
                     "Материал корпуса", "Тип", "Прочее"]
    for product in parser.data:
        product_text = product["text"]
        extracted = extract_with_qwen(product_text)
        for col in final_columns:
            if col not in extracted:
                extracted[col] = "не указано"
        filled_forms.append(extracted)
    df_form = pd.DataFrame(filled_forms, columns=final_columns)
    print("\nЗаполненная форма:")
    print(df_form.to_string(index=False))
    output_file = "output.xlsx"
    append_df_to_excel(output_file, df_form, sheet_name="Sheet1")
    print(f"\nДанные успешно добавлены в файл {output_file}.")
